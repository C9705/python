import time
import pyautogui as pg
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys  
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import os
import pandas as pd
import traceback
def reardata():
    import pandas as pd
    global CTO
    print(os.path.dirname(__file__))
    pathfile=os.path.dirname(__file__)+"\\osAuto.xlsm"
    data = xlrd.open_workbook(pathfile).sheet_by_name('Sheet1')
    CTO=data.cell(0,0).value
    print(CTO)

def work(Address,CTOnum,count,basicName):
    # login
    # 从txt文件读取LOIS 登陆信息，只需在txt文件维护登陆信息即可
    # file = open(r"E:\workspaceprod\cuichuang1\Daily SBB Asynch Check\lois-login-username-pw.txt", "r")  # 修改文件路径
    # login = file.readlines()
    # username = (login[0].strip()).replace('\n', '')
    # pw = (login[1].strip()).replace('\n', '')
    # file.close()

    #临时
    username="cuichuang1"
    pw='970409Cc.'

    # change the download path
    chromeOptions = webdriver.ChromeOptions()
    # prefs = {"download.default_directory": "E:\workspaceprod\qinn1\MTM_Report_Upgrade\init_mtm"} #修改文件路径
    # chromeOptions.add_experimental_option("prefs", prefs)
    browser = webdriver.Chrome(chrome_options=chromeOptions)
    # browser.implicitly_wait(5)
 

    time.sleep(1)
    browser.get(r'http://lois.lenovo.com/lenovo-lois-web/main/main.html')
    # reload webpage(lois)
    time.sleep(3)
    pg.moveTo(x=103, y=67, duration=0.5)
    time.sleep(0.5)
    pg.click()

    # auto enter itcode and password to login
    time.sleep(3)
    pg.moveTo(x=463, y=182, duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(1)

    # 浏览器最大化
    pg.hotkey("alt"," ")
    time.sleep(0.5)
    pg.hotkey("x")
    time.sleep(0.5)
    pg.hotkey("alt","D")
    time.sleep(1)
    pg.typewrite(username, 0.1)
    time.sleep(1)
    pg.moveTo(x=591, y=211, duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(1)
    pg.typewrite(pw, 0.1)
    # time.sleep(1)
    pg.moveTo(x=775, y=264, duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(1)
    pg.press("enter")
    print(browser.title)
    for handle in browser.window_handles:
        browser.switch_to.window(handle)
        if 'lois' in browser.title:
            break
    browser.refresh()
    time.sleep(1)
    # SBB Management
    SBB=browser.find_element_by_xpath("//*[text()='SBB Management']")
    ActionChains(browser).double_click(SBB).perform()
    time.sleep(0.5)
    #Maintain SBB
    browser.find_element_by_xpath("//*[text()='Maintain SBB']").click()
    time.sleep(0.5)
    #Search
    browser.find_element_by_xpath("//*[text()='Search']").click()
    time.sleep(0.5)
    # CTO NUm
    CTONum=browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/input')
    browser.execute_script("arguments[0].value='"+CTO+"'",CTONum)
     # SBB Basic name
    Basic=browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[1]/div[5]/div/div[1]/div[1]/input')
    browser.execute_script("arguments[0].value='"+basicName+"'",Basic)
    # status
    status1=browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[1]/div[4]/div/div[2]/div[1]/input')
    browser.execute_script("arguments[0].value='Initiated'",status1)
    # Search2
    browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[2]/div/div[2]').click()
    time.sleep(60)
    # more 500
    browser.find_element_by_xpath("/html/body/div[3]/div[2]/div/div[3]/div[2]/div[2]/div/div[1]/div[2]/div/div[3]/div/div[12]/div[1]/div/div[1]").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[text()='"+str(count)+"']").click()
    time.sleep(60)
    #select all
    browser.find_element_by_xpath("/html/body/div[3]/div[2]/div/div[3]/div[2]/div[2]/div/div[1]/div[2]/div/div[4]/div/div[1]/div/span").click()
    time.sleep(0.5)
    #view links
    browser.find_element_by_xpath("//*[text()='View Links']").click()
    time.sleep(0.5)
    browser.find_element_by_xpath("/html/body/div[13]/div[1]/div[2]/div[2]/a/span").click()
    time.sleep(5)
    # Search element
    pg.moveTo(223,150,duration=0.5)
    pg.click()
    # element
    pg.moveTo(525,299,duration=0.5)
    pg.click()    
    time.sleep(1)
    pg.typewrite(message='free-dos',interval=0.5)
    pg.moveTo(998,489,duration=0.5)
    pg.click()
    time.sleep(2)
    # search result
    pg.moveTo(385,320,duration=0.5)
    pg.click()
    time.sleep(0.5)
    # confirm
    pg.moveTo(878,576,duration=0.5)
    pg.click()
    print('done')
    # checkout report
    pg.moveTo(590,149,duration=0.5)
    pg.click()
    # show=browser.find_element_by_xpath("/html/body/div[3]/div[2]/div/div[3]/div[2]/div[2]/div/div[1]/div[2]/div/div[3]/div/div[12]/div[1]/input")
    # browser.execute_script("arguments[0].value='500'",show)
    time.sleep(60)
    browser.quit()
   
def deal(url,url2):
    # 获取download下的文件
    files=os.listdir(url)
    file=filter(lambda x: x.find('sbbElement_Spreadsheet_')!=-1,files)
    file=list(file)
    le=len(file)
    if(le==0):
        print('no sbbElement file')
        return False
    print(True)    
    file=list(file)[0]
    wb=xlrd.open_workbook(url+"\\"+file)
    wb.sheet_by_index(0)
    data=wb.sheet_by_index(0).row_values(5)
    #   写入第六行数据
    wb2=load_workbook(url2)
    ws=wb2.active
    ws.delete_cols(4,500)
    for i in range(0,len(data)):#data[i]
        ws.cell(row=6,column=i+1,value=data[i])
    wb2.save(url2)
    time.sleep(3)
    main=r'C:\Users\cuichuang1\Desktop\VM\Common\element_link_tool.py'
    os.system(main)
    # 删除
    if os.path.exists(url+"\\"+file):  # 如果文件存在
    # 删除文件，可使用以下两种方法。
        os.remove(url+"\\"+file)  
    #os.unlink(path)
    else:   
        print('no such file:%s'%url+"\\"+file)    

#upload
def workUpLoad(Address,CTOnum,count,basicName):
    # login
    # 从txt文件读取LOIS 登陆信息，只需在txt文件维护登陆信息即可
    # file = open(r"E:\workspaceprod\cuichuang1\Daily SBB Asynch Check\lois-login-username-pw.txt", "r")  # 修改文件路径
    # login = file.readlines()
    # username = (login[0].strip()).replace('\n', '')
    # pw = (login[1].strip()).replace('\n', '')
    username='cui'
    pw='97'
    # file.close()

    # change the download path
    chromeOptions = webdriver.ChromeOptions()
    # prefs = {"download.default_directory": "E:\workspaceprod\qinn1\MTM_Report_Upgrade\init_mtm"} #修改文件路径
    # chromeOptions.add_experimental_option("prefs", prefs)
    browser = webdriver.Chrome(chrome_options=chromeOptions)
    # browser.implicitly_wait(5)

 


    time.sleep(1)
    browser.get(r'http://lois.lenovo.com/lenovo-lois-web/main/main.html')
    # reload webpage(lois)
    time.sleep(3)
    pg.moveTo(x=103, y=67, duration=0.5)
    time.sleep(0.5)
    pg.click()

    # auto enter itcode and password to login
    time.sleep(3)
    pg.moveTo(x=463, y=182, duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(1)

    # 浏览器最大化
    pg.hotkey("alt"," ")
    time.sleep(0.5)
    pg.hotkey("x")
    time.sleep(0.5)
    pg.hotkey("alt","D")
    time.sleep(1)
    pg.typewrite(username, 0.1)
    time.sleep(1)
    pg.moveTo(x=591, y=211, duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(1)
    pg.typewrite(pw, 0.1)
    # time.sleep(1)
    pg.moveTo(x=775, y=264, duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(1)
    pg.press("enter")
    print(browser.title)
    for handle in browser.window_handles:
        browser.switch_to.window(handle)
        if 'lois' in browser.title:
            break
    browser.refresh()
    time.sleep(1)
    # SBB Management
    SBB=browser.find_element_by_xpath("//*[text()='SBB Management']")
    ActionChains(browser).double_click(SBB).perform()
    time.sleep(0.5)
    #Maintain SBB
    browser.find_element_by_xpath("//*[text()='Maintain SBB']").click()
    time.sleep(0.5)
    #Search
    browser.find_element_by_xpath("//*[text()='Search']").click()
    time.sleep(0.5)
    # CTO NUm
    CTONum=browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/input')
    browser.execute_script("arguments[0].value='"+CTO+"'",CTONum)
     # SBB Basic name
    Basic=browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[1]/div[5]/div/div[1]/div[1]/input')
    browser.execute_script("arguments[0].value='"+basicName+"'",Basic)
    # status
    status1=browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[1]/div[4]/div/div[2]/div[1]/input')
    browser.execute_script("arguments[0].value='Initiated'",status1)
    # Search2
    browser.find_element_by_xpath('/html/body/div[11]/div[2]/div/div[2]/div/div[2]').click()
    time.sleep(60)
    # more 500
    browser.find_element_by_xpath("/html/body/div[3]/div[2]/div/div[3]/div[2]/div[2]/div/div[1]/div[2]/div/div[3]/div/div[12]/div[1]/div/div[1]").click()
    time.sleep(1)
    browser.find_element_by_xpath("//*[text()='"+str(count)+"']").click()
    time.sleep(60)
    #select all
    browser.find_element_by_xpath("/html/body/div[3]/div[2]/div/div[3]/div[2]/div[2]/div/div[1]/div[2]/div/div[4]/div/div[1]/div/span").click()
    time.sleep(1)
    #ELement link loadsheet
    browser.find_element_by_xpath('/html/body/div[3]/div[2]/div/div[3]/div[2]/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/div[7]/em/button/span/b').click()
    time.sleep(1)
    browser.find_element_by_xpath('/html/body/div[13]/div[1]/div[2]/div[3]/a/span').click()
    time.sleep(1)
    #   links window
    pg.moveTo(792,292,duration=0.5)
    pg.click()
    time.sleep(1)
    # file uploadCo
    # pg.moveTo(315,416,duration=0.5)
    # pg.click()
    time.sleep(1)
    # pg.typewrite(message=r"E:\workspaceprod\Common\oslink_database\upload_os.xlsx",interval=0.5)
    pg.typewrite(message=r"C:\Users\cuichuang1\Desktop\VM\Common\oslink_database\upload_os.xlsx",interval=0.5)
    pg.press('enter')
    time.sleep(1)
    pg.moveTo(624,484,duration=0.5)
    time.sleep(1)
    pg.click()
    time.sleep(5)
    pg.moveTo(871,286,duration=0.5)
    time.sleep(10)
    pg.click()
    pg.click()
    time.sleep(30)
    browser.quit()
   


 

if __name__ == "__main__":
    
    import  sys
    f=open(r"C:\Users\cuichuang1\Desktop\VM\cuichuang1\Element_Auto\osX.txt","r+")
    f.truncate()
    f.close()
    class Logger(object):
        def __init__(self, fileN="Default.log"):
            self.terminal = sys.stdout
            self.log = open(fileN, "a+", encoding='utf8')

        def write(self, message):
            self.terminal.write(message)
            self.log.write(message)

        def flush(self):
            pass
    sys.stdout = Logger(r"C:\Users\cuichuang1\Desktop\VM\cuichuang1\Element_Auto\osX.txt")  # 这里我将Log输出到D盘
    sys.stderr = Logger(r"C:\Users\cuichuang1\Desktop\VM\cuichuang1\Element_Auto\osX.log")
    reardata()
    list1=['L1 CA','L1 CD','L1 PD','L1 RHD','L1 VA','L1 BAT','L1 MPA','L1 PCMPA']
    log=[]
    for i in list1:
        try:
            work(r'http://lois.lenovo.com/lenovo-lois-web/main/main.html',CTO,50,i)
            if deal(r'C:\Users\cuichuang1\Downloads',r"C:\Users\cuichuang1\Desktop\VM\Common\oslink_database\upload_os.xlsx") !=False:
                print('normal')
                workUpLoad(r'http://lois.lenovo.com/lenovo-lois-web/main/main.html',CTO,50,i)
        except Exception:
            log.append(traceback.print_exc())
        time.sleep(2)
    # os.popen('taskkill.exe /pid:'+str(pid))
    command = 'taskkill /F /IM chrome.exe'
    command2 = 'taskkill /F /IM chromedriver.exe'
    #比如这里关闭QQ进程
    os.system(command)
    os.system(command2)
    pass