import openpyxl as xl
import xlrd
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np

from openpyxl import Workbook
import xlwt
import os
import time
from openpyxl import load_workbook
import sys
import sched

# schedule = sched.scheduler(time.time, time.sleep)

# def perform_command(cmd,inc):
#     os.system(cmd)
#     print('task')
# def timming_exe(cmd,inc=60):
#     schedule.enter(inc,0,perform_command,(cmd,inc))
#     schedule.run()
# print('show time after 2 seconds:')
# timming_exe('echo %time%',2)


def final(LJ1,LJ2):
    wb3 = xl.load_workbook(LJ2)
    sheet2 = wb3.active
    # sheet2.cell(row=7, column=4).value = "B"
    with xlrd.open_workbook(LJ2) as data:
        table = data.sheet_by_index(0)
        first_col = table.col_values(0)
        row_index = first_col.index(r'ID')
        sbb_list = [(each.replace('[', '|').replace(']', '|').split('|')[2], col_index + 1) for col_index, each in
                    enumerate(table.row_values(row_index)) if each.startswith(("W", "C", "V", "A", "B", "D", "E",
                                                                               "F", "G", "H", "I", "J", "K", "L",
                                                                               "M", "N", "O", "P", "Q", "R", "S",
                                                                               "T", "U", "X", "Y", "Z", "1", "2",
                                                                               "3", "4", "5", "6", "7", "8", "9"),
                                                                              12, 13)]

        wb = xl.load_workbook(LJ1)
        sheet = wb.active
        dic = pd.read_excel(LJ1)
        element_dic = dic.set_index("SBB Name").to_dict()["Element Desc"]
        print(element_dic)
        for key1, value in element_dic.items():
            print(key1)

            for i1 in sbb_list:
                # print(key1)
                if i1[0] == key1:
                    x1 = int(i1[1])
                    x2 = element_dic[key1]

                    print(x1)
                    print(x2)

                    for row_index2 in sheet2["C6":"C{}".format(sheet2.max_row + 1)]:

                        for cell1 in row_index2:
                            # print (cell1.value)
                            if x2 == cell1.value:
                                row_num = cell1.row
                                print(row_num)
                                # print(x1)
                                sheet2.cell(row=row_num, column=x1).value = 'A'
                                # sheet2.cell(row=7,column=4).value="B"
                    # sheet.cell(row=7, column={ }).value = 'A' .format(int(i3[1]))

    wb3.save(LJ2)



# def main():
#     current_dir = os.path.split(os.path.realpath(__file__))[0]
#     filenames = [os.path.join(current_dir,each) for each in os.listdir(current_dir) if os.path.isfile(each) and os.path.splitext(each)[1] in ('.xlsx', '.xls') and not each.startswith('~') and not 'result' in each]
#
#
#
#     filenames2 = [os.path.join(current_dir,each) for each in os.listdir(current_dir)]
#     spreadsheet = [each for each in filenames2 if 'upload' in each.lower()]
#     element = [each for each in filenames2 if 'database' in each.lower()]
#     windchill = [each for each in filenames2 if 'windchill information' in each.lower()]
#     print (spreadsheet)
#     # print 'spreadsheet:{0}\nelement:{1}\nwindchill:{2}'.format(spreadsheet, element, windchill)
#     for each_spreadsheet in spreadsheet:
#         print (each_spreadsheet)
#         for each_element in element:
#             final(each_element, each_spreadsheet)

if __name__ == '__main__':
    print (u'''***************************************************************************************************
please don't set this tool in Chinese path and don't stop the process during it running.thanks
***************************************************************************************************''')
    # final(r'E:\workspaceprod\Common\oslink_database\database_os.xlsx',r'E:\workspaceprod\Common\oslink_database\upload_os.xlsx')
    final(r'C:\Users\cuichuang1\Desktop\VM\Common\oslink_database\database_os.xlsx',r'C:\Users\cuichuang1\Desktop\VM\Common\oslink_database\upload_os.xlsx')

print('When you see me, it proves that everything is done O(∩_∩)O')
